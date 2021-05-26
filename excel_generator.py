from datetime import datetime
from io import BytesIO
from urllib.parse import quote_plus

from aiohttp import web
from openpyxl.utils import get_column_letter
from pytz import timezone
from xlrd import open_workbook

from .excel_utils import Workbook, Style
from .loggers import getLogger
from .queries import get_object
from .settings import get_config
from .schema_validators import ValidError

logger = getLogger()


class DataBaseKeys:
    """Ключи полей в БД."""


class ReportBase:
    CONFIG = get_config()['reports']


class ReportFormat:
    EXCEL = 'xlsx'
    EXCEL_2003 = 'xls'
    DEFAULT = EXCEL

    # Словарь форматов данных
    ALL = {
        EXCEL: 'Excel 2007+',
        EXCEL_2003: 'Excel 2003',
    }
    ALL_LIST = [{'id': k, 'name': v} for k, v in ALL.items()]

    # Тип файла отчета
    FILE_TYPE = {
        EXCEL: '.xlsx',
        EXCEL_2003: '.xls',
    }


class ExcelCell:
    CELL = 'cell'
    CELL_BIG = 'cell_big'
    CELL_PERCENT = 'cell_percent'
    CELL_DATE = 'cell_date'
    CELL_DATETIME = 'cell_datetime'
    HEADER = 'header'
    HEADER_SLIM = 'header_slim'


class ExcelRow(ExcelCell):
    # Вертикальные отсупы в строках
    _ADDITIONAL_HEIGHT_CELL = 2
    _ADDITIONAL_HEIGHT_HEADER = 1

    HEIGHTS = {
        ExcelCell.CELL: _ADDITIONAL_HEIGHT_CELL,
        ExcelCell.HEADER: _ADDITIONAL_HEIGHT_HEADER,
    }


class ExcelColumn(ExcelCell):
    # Горизонтальные отсупы в колонках
    _ADDITIONAL_WIDTH_CELL = 7
    _ADDITIONAL_WIDTH_CELL_BIG = 31
    _ADDITIONAL_WIDTH_CELL_SLIM = 3

    WIDTHS = {
        ExcelCell.CELL: _ADDITIONAL_WIDTH_CELL,
        ExcelCell.CELL_BIG: _ADDITIONAL_WIDTH_CELL_BIG,
        ExcelCell.HEADER_SLIM: _ADDITIONAL_WIDTH_CELL_SLIM,
    }


class ExcelCellStyle:
    def __init__(self, header, cell, column: str=ExcelColumn.CELL,
                    width: int=0):
        self._header = header
        self._cell = cell
        self._column = column
        self._width = width + ExcelColumn.WIDTHS[column]

    @property
    def column(self):
        return self._column

    @property
    def width(self):
        return self._width

    @property
    def header(self):
        return self._header

    @property
    def cell(self):
        return self._cell


class ExcelSheetBase(ExcelRow, ExcelColumn):
    # Директория хранения таблиц
    DIR                         = ReportBase.CONFIG['dir']
    # Ключи ячеек
    # Необходимые ключи

    # Дополнительные Ключи
    # Номер строки
    ROW_NUM                     = ''
    # Цвет строки
    COLORED                     = '_colored'
    # Наименование ключа итогового результата счетчика
    TOTAL                       = 'total'
    PERCENT                     = 'percent'
    # Типы таблиц отчетов
    REPORT                      = 'report'
    STATISTIC                   = 'statistic'
    # Таблицы
    REPORT_SHORT                = REPORT + '_short'
    REPORT_FULL                 = REPORT + '_full'
    STATISTIC_REPORT            = STATISTIC + '_'
    # Архитектура таблиц {'наименование стобца': 'ключ в массиве данных'}:
    _REPORT_SHORT = {
        '№':                    ROW_NUM,
        # Столбцы
    }
    _REPORT_FULL = {
        '№':                    ROW_NUM,
        # Столбцы
    }
    _STATISTIC = {
        '№':                    ROW_NUM,
        # Столбцы
    }
    # Сортировка таблиц
    SORT = {
        REPORT_SHORT: lambda v: (
            # поля
        ),
        REPORT_FULL: lambda v: (
            # поля
        ),
    }
    # Словарь данных таблиц
    ALL = {
        REPORT_SHORT: {
            'sheet':            _REPORT_SHORT,
            'sheet_name':       'Краткий отчет',
            'sheet_full':       REPORT_FULL,
        },
        REPORT_FULL: {
            'sheet':            _REPORT_FULL,
            'sheet_name':       'Расширенный отчет',
        },
        STATISTIC_REPORT: {
            'sheet':            _STATISTIC,
            'sheet_name':       'Статистика',
        },
    }

    @classmethod
    def _optimise_sheets_list(cls, sheets_list: list) -> list:
        if not sheets_list:
            raise ValidError.field_is_empty('sheets', 'Список таблиц')

        if (cls.REPORT_FULL in sheets_list):
            sheets_list = [
                v for v in set(sheets_list) if v != cls.REPORT_SHORT
            ]

        return sheets_list

    @classmethod
    def __get_sheet(cls, sheet_key: str) -> dict:
        try:
            return cls.ALL[sheet_key]
        except KeyError:
            raise ValidError.field_incorrect_value(
                'sheet_key', 'Таблица', sheet_key)

    @classmethod
    def _get_sheet(cls, sheet_key: str) -> dict:
        _sheet = cls.__get_sheet(sheet_key)
        return _sheet['sheet']

    @classmethod
    def _get_sheet_name(cls, sheet_key: str) -> str:
        _sheet = cls.__get_sheet(sheet_key)
        return _sheet['sheet_name']

    @classmethod
    def _has_field(cls, sheet_key: str, field: str) -> bool:
        """Проверка на наличие поля в Таблице."""
        return field in cls._get_sheet(sheet_key).values()


class Report(ReportBase):
    # Пользовательский отчет
    REPORT_CUSTOM = ExcelSheetBase.REPORT + '_custom'
    # Короткий отчет
    REPORT_SHORT = ExcelSheetBase.REPORT_SHORT
    # Полный отчет
    REPORT_FULL = ExcelSheetBase.REPORT_FULL
    # Статистический отчет
    STATISTIC = ExcelSheetBase.STATISTIC
    # Отчет ЕДГ для отправки в СС
    REPORT_SHORT_EDG = ReportBase._get_edg_key(REPORT_SHORT)
    # Отчет СМК для отправки в СС
    REPORT_SHORT_SMK = ReportBase._get_smk_key(REPORT_SHORT)
    # Полный отчёт ЕДГ
    REPORT_FULL_EDG = ReportBase._get_edg_key(REPORT_FULL)
    # Полный отчёт СМК
    REPORT_FULL_SMK = ReportBase._get_smk_key(REPORT_FULL)
    # Отчёт-статистика ЕДГ
    STATISTIC_EDG = ReportBase._get_edg_key(STATISTIC)
    # Отчёт-статистика СМК
    STATISTIC_SMK = ReportBase._get_smk_key(STATISTIC)
    # Отчет по умолчанию
    DEFAULT = REPORT_CUSTOM
    # Словарь отчетов
    ALL = {
        REPORT_CUSTOM:      'Пользовательский',
        REPORT_SHORT:       'Список подозрений для отправки в СС',
        REPORT_FULL:        'Расширенный список подозрений',
        STATISTIC:          'Статистика подозрений',
    }
    ALL_LIST = [{'id': k, 'name': v} for k, v in ALL.items()]
    ALL_REPORTS = {
        REPORT_CUSTOM:      'Пользовательский отчёт',
        REPORT_SHORT_EDG:   'Список подозрений ЕДГ для отправки в СС за период',
        REPORT_SHORT_SMK:   'Список подозрений СМК для отправки в СС за период',
        REPORT_FULL_EDG:    'Расширенный список подозрений ЕДГ за период',
        REPORT_FULL_SMK:    'Расширенный список подозрений СМК за период',
        STATISTIC_EDG:      'Статистика подозрений ЕДГ за период',
        STATISTIC_SMK:      'Статистика подозрений СМК за период',
    }
    TYPES = {
        REPORT_CUSTOM:      [],
        REPORT_SHORT_EDG:   [
            ExcelSheetBase.REPORT_SHORT,
        ],
        REPORT_SHORT_SMK:   [
            ExcelSheetBase.REPORT_SHORT,
        ],
        REPORT_FULL_EDG:    [
            ExcelSheetBase.REPORT_FULL,
        ],
        REPORT_FULL_SMK:    [
            ExcelSheetBase.REPORT_FULL,
        ],
        STATISTIC_EDG:      [
            ExcelSheetBase.STATISTIC_SOURCES,
            ExcelSheetBase.STATISTIC_CATEGORIES,
            ExcelSheetBase.STATISTIC_SOURCE_TYPES,
            ExcelSheetBase.STATISTIC_INFORMATION_TYPES,
            ExcelSheetBase.STATISTIC_TIMINGS,
            ExcelSheetBase.STATISTIC_LANGUAGES,
            ExcelSheetBase.STATISTIC_FSEM,
        ],
        STATISTIC_SMK:      [
            ExcelSheetBase.STATISTIC_SOURCES,
            ExcelSheetBase.STATISTIC_CATEGORIES,
            ExcelSheetBase.STATISTIC_SOURCE_TYPES,
            ExcelSheetBase.STATISTIC_INFORMATION_TYPES,
            ExcelSheetBase.STATISTIC_TIMINGS,
            ExcelSheetBase.STATISTIC_LANGUAGES,
        ],
    }

    @classmethod
    def get_type(cls, report_type: str) -> str:
        if not report_type or report_type == cls.REPORT_CUSTOM:
            return cls.REPORT_CUSTOM
        else:
            return report_type


class ExcelColors:
    GRAY = 'gray'
    GREEN = 'green'
    RED = 'red'
    YELLOW = 'yellow'


class ExcelSheetStyleBase(ExcelColors):
    _STATUS_STYLE = 'default_status_style'
    _TOTAL_STYLE = 'default_total_style'
    _PERCENT_STYLE = 'default_percent_style'
    _TOTAL = ExcelSheetBase.TOTAL
    _PERCENT = ExcelSheetBase.PERCENT
    STYLES = {}

    def get_style(self, column: str) -> ExcelCellStyle:
        try:
            return self.STYLES[column]
        except KeyError:
            if self._PERCENT in column:
                return self.STYLES[self._PERCENT_STYLE]
            elif self._TOTAL in column:
                return self.STYLES[self._TOTAL_STYLE]
            else:
                return self.STYLES[self._STATUS_STYLE]


class ExcelSheetStyle(ExcelSheetStyleBase):
    ALL = lambda    header_style, header_slim_style, cell_style,\
                    cell_big_style, cell_percent_style, cell_date_style,\
                    cell_datetime_style: {
        ExcelSheetBase.HEADER:                  header_style,
        ExcelSheetBase.HEADER_SLIM:             header_slim_style,
        ExcelSheetBase.CELL:                    cell_style,
        ExcelSheetBase.CELL_BIG:                cell_big_style,
        ExcelSheetBase.CELL_PERCENT:            cell_percent_style,
        ExcelSheetBase.CELL_DATE:               cell_date_style,
        ExcelSheetBase.CELL_DATETIME:           cell_datetime_style,
    }

    STYLES = lambda header_style, header_slim_style, cell_style,\
                    cell_big_style, cell_percent_style, cell_date_style: {
        ExcelSheetBase.ROW_NUM:                 ExcelCellStyle(
            header=header_style,
            cell=cell_style,
        ),
        ExcelSheetBase.PUBLICATION_DATETIME:    ExcelCellStyle(
            header=header_style,
            cell=cell_date_style,
        ),
        ExcelSheetBase.CREATED_AT:              ExcelCellStyle(
            header=header_style,
            cell=cell_date_style,
        ),
        ExcelSheetBase.PROCESSED_AT:            ExcelCellStyle(
            header=header_style,
            cell=cell_date_style,
        ),
        ExcelSheetBase.SOURCE:                  ExcelCellStyle(
            header=header_style,
            cell=cell_style,
        ),
        ExcelSheetBase.STATUS:                  ExcelCellStyle(
            header=header_style,
            cell=cell_style,
        ),
        ExcelSheetBase.REGISTRY_STATUS:         ExcelCellStyle(
            header=header_style,
            cell=cell_style,
        ),
        ExcelSheetBase.SENT_VIOLATION:          ExcelCellStyle(
            header=header_style,
            cell=cell_style,
        ),
        ExcelSheetBase.REQUIREMENT:             ExcelCellStyle(
            header=header_style,
            cell=cell_style,
        ),
        ExcelSheetBase.CATEGORIES:              ExcelCellStyle(
            header=header_style,
            cell=cell_big_style,
            column=ExcelSheetBase.CELL_BIG,
        ),
        ExcelSheetBase.URL:                     ExcelCellStyle(
            header=header_style,
            cell=cell_big_style,
            column=ExcelSheetBase.CELL_BIG,
        ),
        ExcelSheetBase.SOURCE_TYPE:             ExcelCellStyle(
            header=header_style,
            cell=cell_style,
        ),
        ExcelSheetBase.INFORMATION_ACCESS:      ExcelCellStyle(
            header=header_style,
            cell=cell_style,
        ),
        ExcelSheetBase.INFORMATION_TYPES:       ExcelCellStyle(
            header=header_style,
            cell=cell_style,
        ),
        ExcelSheetBase.TIMING:                  ExcelCellStyle(
            header=header_style,
            cell=cell_style,
        ),
        ExcelSheetBase.LANGUAGES:               ExcelCellStyle(
            header=header_style,
            cell=cell_style,
        ),
        ExcelSheetBase.FSEM:                    ExcelCellStyle(
            header=header_style,
            cell=cell_big_style,
            column=ExcelSheetBase.CELL_BIG,
        ),
        ExcelSheetBase.FEDERAL_DISTRICT:        ExcelCellStyle(
            header=header_style,
            cell=cell_style,
        ),
        ExcelSheetBase.REGION:                  ExcelCellStyle(
            header=header_style,
            cell=cell_style,
        ),
        ExcelSheetBase.USER:                    ExcelCellStyle(
            header=header_style,
            cell=cell_style,
        ),
        ExcelSheetBase.DESCRIPTION:             ExcelCellStyle(
            header=header_style,
            cell=cell_big_style,
            column=ExcelSheetBase.CELL_BIG,
        ),
        # ExcelSheetBase.SOURCE_SYSTEM:           ExcelCellStyle(
        #     header=header_style,
        #     cell=cell_style,
        # ),
        ExcelSheetBase.IS_MANUAL:               ExcelCellStyle(
            header=header_style,
            cell=cell_style,
        ),
        ExcelSheetStyleBase._STATUS_STYLE:      ExcelCellStyle(
            header=header_slim_style,
            cell=cell_style,
            column=ExcelSheetBase.HEADER_SLIM,
        ),
        ExcelSheetStyleBase._PERCENT_STYLE:     ExcelCellStyle(
            header=header_slim_style,
            cell=cell_percent_style,
            column=ExcelSheetBase.HEADER_SLIM,
        ),
        ExcelSheetStyleBase._TOTAL_STYLE:       ExcelCellStyle(
            header=header_slim_style,
            cell=cell_style,
            column=ExcelSheetBase.HEADER_SLIM,
        ),
    }

    COLORS = lambda gray, green, red, yellow: {
        ExcelColors.GRAY:                       gray,
        ExcelColors.GREEN:                      green,
        ExcelColors.RED:                        red,
        ExcelColors.YELLOW:                     yellow,
    }


class ExcelSheet(ExcelSheetStyleBase):
    STYLE = Style

    # Словарь стилей книги
    HEADER_STYLE = STYLE(
        ExcelSheetBase.HEADER,
        font_size=14,
        font_bold=True,
        cell_format=STYLE.Format.TEXT,
        pattern_type=STYLE.Pattern.SOLID,
        pattern_fg_color=STYLE.Color.SLATE_BLUE,
        border_b=STYLE.Border.MEDIUM, border_t=STYLE.Border.MEDIUM,
        border_r=STYLE.Border.MEDIUM, border_l=STYLE.Border.MEDIUM
    ).get_style()
    HEADER_SLIM_STYLE = STYLE(
        ExcelSheetBase.HEADER_SLIM,
        style=HEADER_STYLE,
        align_wrap=True,
    ).get_style()
    CELL_STYLE = STYLE(
        ExcelSheetBase.CELL,
        font_size=12,
        align_wrap=True,
        cell_format=STYLE.Format.TEXT,
        pattern_type=STYLE.Pattern.SOLID,
        border_b=STYLE.Border.THIN, border_t=STYLE.Border.THIN,
        border_r=STYLE.Border.THIN, border_l=STYLE.Border.THIN
    ).get_style()
    CELL_BIG_STYLE = STYLE(
        ExcelSheetBase.CELL_BIG,
        style=CELL_STYLE,
        align_h=STYLE.Alignment.LEFT,
    ).get_style()
    CELL_PERCENT_STYLE = STYLE(
        ExcelSheetBase.CELL_PERCENT,
        style=CELL_STYLE,
        number_format=STYLE.Format.PERCENTAGE_00,
    ).get_style()
    CELL_DATE_STYLE = STYLE(
        ExcelSheetBase.CELL_DATE,
        style=CELL_STYLE,
        number_format=STYLE.Format.DATE,
    ).get_style()
    CELL_DATETIME_STYLE = STYLE(
        ExcelSheetBase.CELL_DATETIME,
        style=CELL_STYLE,
        number_format=STYLE.Format.DATETIME,
    ).get_style()

    ALL = ExcelSheetStyle.ALL(
        HEADER_STYLE, HEADER_SLIM_STYLE, CELL_STYLE, CELL_BIG_STYLE,
        CELL_PERCENT_STYLE, CELL_DATE_STYLE, CELL_DATETIME_STYLE
    )
    STYLES = ExcelSheetStyle.STYLES(
        HEADER_STYLE, HEADER_SLIM_STYLE, CELL_STYLE, CELL_BIG_STYLE,
        CELL_PERCENT_STYLE, CELL_DATE_STYLE
    )
    COLORS = ExcelSheetStyle.COLORS(
        gray=STYLE.Color.GRAY,
        green=STYLE.Color.GREEN,
        red=STYLE.Color.RED,
        yellow=STYLE.Color.YELLOW,
    )

    def get_height(self, row: str=ExcelRow.CELL) -> int:
        return ((self.ALL[row].font.size + ExcelRow.HEIGHTS[row])
                * self.STYLE.FONT_HEIGHT_INC)

    def get_width(self, column: str) -> int:
        cell = self.get_style(column)
        return (cell.header.font.size / 10) * (cell.width + len(column))


class ExcelGenerator(ExcelSheetBase):
    """Генератор excel-файла.

    Args:
        violation_form (str): Форма нарушений в соответствии с ViolationForm.
        data (dict, optional): Данные для выгрузки в excel.
        report_type (int, optional): Тип отчета в соответствии с
            Report (по умолчанию - Report.DEFAULT).
        report_format (int, optional): Формат отчета в соответствии с
            ReportFormat (по умолчанию - ReportFormat.DEFAULT).
        fd (BytesIO, optional): Файловый дискриптор для
            сохранения файла локально.

    """
    # Excel нулевой символ
    NULL_SYMB = None
    NULL_SYMB_TO_CELL = '-'
    # Excel разделитель списка
    DELIMITER = ";\n"
    # Словарь наименований книг
    WB_NAMES = Report.ALL_REPORTS
    DEFAULT_WB_NAME = Report.ALL_REPORTS[Report.DEFAULT]
    # Стили таблиц
    EXCEL_SHEET = ExcelSheet()

    def __init__(self,
        violation_form: str, data: dict = {},
        report_type: int = Report.DEFAULT,
        report_format: int = ReportFormat.DEFAULT,
        fd: BytesIO = None
    ):
        self.__violation_form = violation_form
        self._data = data
        self.report_type = report_type
        self.report_format = report_format
        # file descriptor
        self.__fd = fd
        # Excel-книга
        self.__wb = None
        # Имя файла/книги
        self.__fn = None
        self.__response = None

    @property
    def violation_form(self) -> str:
        return self.__violation_form

    @property
    def response(self) -> web.Response:
        return self.__response

    @property
    def wb(self) -> Workbook:
        return self.__wb

    @property
    def fd(self) -> BytesIO:
        # если fd не указан -> сохранение excel-книги в буфер
        self.__fd = self.__fd or BytesIO()
        return self.__fd

    @property
    def fn(self) -> str:
        return self.__fn

    def get_sheet(self, sheet_key: str) -> dict:
        return self._get_sheet(self.violation_form, sheet_key)

    def get_sheet_name(self, sheet_key: str) -> str:
        return self._get_sheet_name(self.violation_form, sheet_key)

    def get_sheet_data(self, sheet_key: str) -> list:
        return self._data.get(sheet_key, [])

    def has_field(self, sheet_key: str, field: str) -> bool:
        return self._has_field(self.violation_form, sheet_key, field)

    def _to_xlsx(self):
        if self.report_format != ReportFormat.EXCEL:
            self.save_excel_to_fd()
            wb = open_workbook(file_contents=self.fd.getvalue())
            workbook = Workbook()

            for i in range(0, wb.nsheets):
                sh = wb.sheet_by_index(i)
                sheet = workbook.active if i == 0 else workbook.create_sheet()
                sheet.title = sh.name

                for row in range(0, sh.nrows):
                    for col in range(0, sh.ncols):
                        sheet.cell(row=row + 1,
                                column=col + 1).value = sh.cell_value(row, col)

            self.__wb = workbook

    def save_excel(self):
        """Сохранение сгенерированной ранее книги
        в директорию SHEETS_DIR.
        """
        if self.fn is None:
            errText = 'Error: Workbook does not exist. Genereate it first.'
            logger.error(errText)
            raise web.HTTPInternalServerError(text=errText)

        self.wb.save(ExcelSheetBase.DIR + self.fn)

        logger.debug(
            f'Saving Excel done: {datetime.now().strftime("%H:%M:%S")}')

    def save_excel_to_fd(self, fd=None):
        """Выгрузка сгенерированной ранее книги в буфер
        файлового дискриптора fd.

        Args:
            fd (optional, any): Файловый дискриптор для выгрузки книги.

        """
        if self.fn is None:
            errText = 'Error: Workbook does not exist. Genereate it first.'
            logger.error(errText)
            raise web.HTTPInternalServerError(text=errText)

        if fd is not None:
            self.__fd = fd

        self.wb.save(self.fd)

    def generate_response(self) -> web.Response:
        """Формирование ответа с вложением excel-книги.

        Returns:
            Response: Сформированный ответ библиотеки aiohttp
                с вложенной excel-книгой.

        """
        self.save_excel_to_fd()

        # Удаляем сформированную книгу
        self.__wb = None

        # Формируем ответ с вложенной excel-книгой
        file_name = quote_plus(self.fn)
        self.__response = web.Response(
            body=self.fd.getvalue(),
            headers={'Content-Disposition': f'attachment;filename={file_name}'},
            content_type='application/vnd.ms-excel',
        )
        logger.debug(
            f'Generating response done: {datetime.now().strftime("%H:%M:%S")}')

        return self.__response

    def get_response(self) -> web.Response:
        """Генерация excel-книги и формирование ответа с её вложением.

        Returns:
            Response: Сформированный ответ библиотеки aiohttp
                с вложенной excel-книгой.

        """
        self.generate_excel()
        return self.generate_response()

    def get_excel(self):
        """Генерация excel-книги с последующим локальным сохранением.
        """
        self.generate_excel()
        self.save_excel()

    def generate_excel(self):
        """Генерация excel-книги в зависимости от
        требуемого формата отчета.
        """
        self.generate_excel_xlsx()

        # Чистим сформированный массив данных
        self._data = {}

    @staticmethod
    def _format_wb_name(wb_name: str, report_format: str) -> str:
        return (
            f'{wb_name.replace(" ", "_")}__'
            f'{datetime.now().strftime("%d_%m_%Y")}'
            f'{ReportFormat.FILE_TYPE[report_format]}'
        )

    def generate_excel_xlsx(self):
        """Генерация книги excel библиотеки openpyxl.
        """
        # --------------Формирование excel-книги--------------
        # Префекс наименования книги
        wbName = self.WB_NAMES.get(self.report_type, self.DEFAULT_WB_NAME)
        # Имя excel-файла: wbName__дата.EXCEL_FILE_TYPE
        self.__fn = self._format_wb_name(wbName, ReportFormat.EXCEL)
        # Создаём excel-книгу
        self.__wb = Workbook()

        # Удаляем дефолтную таблицу
        self.__wb.remove_sheet(self.__wb.active)

        # Генерация таблиц книги
        for sheet_key in self._data.keys():
            self._generate_excel_sheet_xlsx(sheet_key)

    def _generate_excel_sheet_xlsx(self, sheet_key: str):
        """Генерация таблицы excel библиотеки openpyxl.
        """
        # Создаем страницу в книге кастомного класса ячеек
        sheet_name = self.get_sheet_name(sheet_key)
        sh = self.__wb.create_sheet(title = sheet_name)

        # Таблица параметров
        SheetStyle = self.EXCEL_SHEET
        sheet = self.get_sheet(sheet_key)

        # Шапка таблицы
        sh_header = list(sheet.keys())
        # Ключи словаря данных, используемых в таблице
        sh_keys = list(sheet.values())

        # ------------------Параметры таблицы------------------
        # Начальный столбец таблицы
        sh_start_column = 1
        # Конечный столбец таблицы
        sh_end_column = len(sh_header) + sh_start_column
        # Начальная строка таблицы
        sh_start_row = 1

        # ------------------Заполнение таблицы------------------
        # Параметры:
        #   rn - row number,
        #   rc - row counter,
        #   cn - column number,
        #   cc - column counter.
        rn = sh_start_row

        logger.debug(f'Excel generation: {datetime.now().strftime("%H:%M:%S")}')

        # Предварительно заполним и настроим таблицу:
        for cc, cn in enumerate(range(sh_start_column, sh_end_column)):
            column = sh_keys[cc]
            column_letter = get_column_letter(cn)
            column_obj = sh.column_dimensions[column_letter]

            # Установим ширину ячеек в соответстии с длиной заголовка:
            width = SheetStyle.get_width(column)
            column_obj.width = width

            # Заполним шапку таблицы:
            sh.cell(rn, cn, sh_header[cc], SheetStyle.get_style(column).header)
            try:
                self.wb.add_named_style(SheetStyle.get_style(column).cell)
            except ValueError:
                pass

        # Заполним таблицу:
        sheet_data = self.get_sheet_data(sheet_key)
        default_style = SheetStyle.CELL_STYLE

        # 1. Заполним нумерованный столбец:
        rc = 1
        column = sh_keys[0]
        try:
            for rn, d in enumerate(sheet_data, start = sh_start_row + 1):
                cell_data = d[column]

                # Если столбец - нумерованный список данных:
                if cell_data is None:
                    cell = sh.cell(rn, sh_start_column, rc, default_style.name)
                    cell.number_format = Style.Format.NUMBER
                    rc += 1
                else:
                    cell = sh.cell(rn, sh_start_column, cell_data,
                                    SheetStyle.get_style(column).cell.name)

        except KeyError:
            for rn, d in enumerate(sheet_data, start = sh_start_row + 1):
                cell = sh.cell(rn, sh_start_column, rc, default_style.name)
                cell.number_format = Style.Format.NUMBER
                rc += 1

        logger.debug(
            f'Excel preparing done: {datetime.now().strftime("%H:%M:%S")}')

        # 2. Заполним столбцы с данными:
        for cc, cn in enumerate(
            range(sh_start_column + 1, sh_end_column), start = 1
        ):
            column = sh_keys[cc]
            column_style = SheetStyle.get_style(column).cell

            for rn, d in enumerate(sheet_data, start = sh_start_row + 1):
                cell_data = d[column]

                if cell_data is None:
                    cell_data = self.NULL_SYMB_TO_CELL
                    cell_style = default_style
                else:
                    cell_style = column_style

                # Если есть цвет строки, устанавливаем новый стиль:
                color = d.get(self.COLORED)
                if color:
                    try:
                        style_name = f'{cell_style.name}_{color}'
                        cell_style = self.wb._named_styles[style_name]
                    except KeyError:
                        cell_style = SheetStyle.STYLE(
                            style_name,
                            style = cell_style,
                            pattern_fg_color = SheetStyle.COLORS[color],
                        ).get_style()
                        self.wb.add_named_style(cell_style)

                cell = sh.cell(rn, cn, cell_data, cell_style.name)

        logger.debug(
            f'Excel table is ready: {datetime.now().strftime("%H:%M:%S")}')


class BaseGenerator(DataBaseKeys):
    def __init__(self, app: dict, query, **kwargs):
        self.__app = app
        self._query = query

        self._data = {}

    @property
    def data(self) -> dict:
        return self._data

    @property
    def app(self) -> dict:
        return self.__app

    @property
    def engine(self):
        return self.__app['db']

    @property
    def query(self):
        return self._query

    # Необходимые аттрибуты

    @classmethod
    def bool_to_str(cls, value: bool) -> str:
        return '+' if value else cls.NULL_SYMB

    async def get_obj(self, to_dict: bool = True):
        """Получаем списки исходных данных.
        """
        if not hasattr(self, 'obj_list'):
            # Получаем список данных из локальной БД
            async with self.engine.acquire() as conn:
                self.obj_list = await get_object(
                    conn, self.query, True)

            self.obj = {
                v['id']: v for v in self.obj_list
            }

        return self.obj if to_dict else self.obj_list

    # Определяем методы
    # ...


class ReportGenerator(BaseGenerator, ExcelGenerator, ReportBase):
    """Генератор отчёта.

    Args:
        app (dict): Текущее приложение.
        query (Query): Параметры запроса.
        sheets_list (list, optional): Список таблиц.
        report_type (int, optional): Тип отчета в соответствии с
            Report (по умолчанию - Report.DEFAULT).
        report_format (int, optional): Формат отчета в соответствии с
            ReportFormat (по умолчанию - ReportFormat.DEFAULT).
        time_column (bool, optional): Признак столбца времени
            (по умолчанию - False).
        fd (BytesIO, optional): Файловый дискриптор для
            сохранения файла локально.

    """
    # Цвета ячеек
    COLORS = ExcelColors
    _COLOR_PRE_TOTAL = COLORS.YELLOW
    _COLOR_TOTAL = COLORS.GREEN

    # Типы отчетов
    REPORTS = Report.TYPES

    def __init__(self,
        app: dict, query,
        sheets_list: list = [],
        report_type: int = Report.DEFAULT,
        report_format: int = ReportFormat.DEFAULT,
        time_column: bool = False,
        fd: BytesIO = None,
        **kwargs
    ):
        ExcelGenerator.__init__(self,
            report_type=report_type,
            report_format=report_format,
            fd=fd,
            **kwargs
        )
        super().__init__(
            app=app,
            query=query,
            **kwargs
        )
        self.sheets_list = sheets_list
        self.time_column = time_column

    # Переопределяем методы, если требуется
    # ...

    #################### Генераторы списков данных ######################
    async def _generate_report(self, sheet_key: str) -> list:
        # Объявляем результирующий список отчётных данных
        # для формирования списка объектов:
        data_list = []

        # Параметры:
        # Московская временная зона для приведения времени
        msc_timezone = timezone('Europe/Moscow')
        # Разделитель многострочного параметра
        delimiter = ExcelGenerator.DELIMITER
        # Нулевой символ (заполнитель пустого значения)
        null_symb = ExcelGenerator.NULL_SYMB

        ####################### Получаем данные #######################
        # Получаем списки идентификаторов
        logger.debug(
            f'Start generate data: {datetime.now().strftime("%H:%M:%S")}')

        # Получаем список данных
        await self.get_obj()
        logger.debug(f'Obj done: {datetime.now().strftime("%H:%M:%S")}')
        # Получаем оставшиеся данные
        # ...

        ################## Формируем список данных ##################
        for o in self.obj_list:
            ################## Формируем словарь данных ##################
            d = {}

            """Определяем порядок заполнения словаря, например:
            # Языки, используемые в тексте материала
            try:
                languages = s[self.DB_LANGUAGES]
                d[self.LANGUAGES] = (
                    delimiter.join([
                        self.languages[v] for v in languages
                    ])
                    if languages else null_symb
                )
            except KeyError as err:
                raise ValidError.obj_data_not_exist('Объект', s[self.DB_ID],
                                                'Язык', self.DB_LANGUAGES, err)

            # Дата создания
            d[self.CREATED_AT] = datetime.fromtimestamp(
                d[self.CREATION_TS], msc_timezone)

            """

            data_list.append(d)

        # Если установлен признак времени,
        # меняем формат поля даты:
        if self.time_column:
            self.EXCEL_SHEET.STYLES = ExcelSheetStyle.STYLES(
                self.EXCEL_SHEET.HEADER_STYLE,
                self.EXCEL_SHEET.HEADER_SLIM_STYLE,
                self.EXCEL_SHEET.CELL_STYLE,
                self.EXCEL_SHEET.CELL_BIG_STYLE,
                self.EXCEL_SHEET.CELL_PERCENT_STYLE,
                self.EXCEL_SHEET.CELL_DATETIME_STYLE
            )
        else:
            self.EXCEL_SHEET.STYLES = ExcelSheet.STYLES

        # Сортируем данные:
        data_list = sorted(data_list, key=self.SORT[sheet_key])
        logger.debug(f'Report data count={len(data_list)} is ready: '
                    f'{datetime.now().strftime("%H:%M:%S")}')

        return data_list

    async def generate_data(self, sheets_list: list=[]) -> dict:
        if sheets_list:
            self.sheets_list = sheets_list

        _sheets_list = (self.sheets_list
                        + self.REPORTS.get(self.report_type, []))
        self.sheets_list = self._optimise_sheets_list(_sheets_list)

        for sheet_key in self.sheets_list:
            if sheet_key in (self.REPORT_SHORT, self.REPORT_FULL):
                self._data[sheet_key] = await self._generate_report(sheet_key)

            elif sheet_key == self.STATISTIC_SOURCES:
                self._data[sheet_key] = await self._generate_statistic_source()

            elif sheet_key == self.STATISTIC_TIMINGS:
                self._data[sheet_key] = await self._generate_statistic_timings()

            else:
                self._data[sheet_key] = await self._generate_statistic_data(
                                                                    sheet_key)

        return self._data


class Statistic:
    GENERAL = 'general'

    DEFAULT = GENERAL

    ALL = {
        GENERAL: 'Основная статистика Подозрений',
    }
    ALL_LIST = [{'id': k, 'name': v} for k, v in ALL.items()]

    _DEFAULT_FIELDS = [
        # Поля ждя ускорения времени генерации
        # Например:
        # table.c.id,
    ]
    FIELDS = {
        GENERAL: _DEFAULT_FIELDS,
        # ...
    }

    @classmethod
    def get_fields(cls, statistic_type: str) -> list:
        return cls.FIELDS.get(statistic_type, [])


class StatisticGenerator(BaseGenerator, Statistic):
    COUNT = 'count'
    OTHERS = 'others'
    STATUSES = 'statuses'

    def __init__(self,
        app: dict, query,
        statistic_type: str = Statistic.DEFAULT,
        **kwargs
    ):
        super().__init__(
            app=app,
            query=query,
            **kwargs
        )
        self._statistic_type = statistic_type

    @property
    def statistic_type(self):
        return self._statistic_type

    # Переопределяем методы, если необходимо:
    # ...

    def generate_response(self) -> web.Response:
        """Формирование ответа.

        Returns:
            Response: Сформированный ответ библиотеки aiohttp.

        """
        self.__response = web.json_response(self._data)
        return self.__response

    async def get_response(self) -> web.Response:
        """Генерация статистических данных и формирование ответа.

        Returns:
            Response: Сформированный ответ библиотеки aiohttp.

        """
        await self.generate_statistic()
        return self.generate_response()

    """Определяем необходимые методы
    Например:

    def __get_src_data(self, key: str) -> dict:
        if key == self.SOURCES:
            return self.sources
        else:
            return self.social_nets

    def __get_data_counter(self, src_data: dict) -> dict:
        # Единичный элемент счетчика в разрезе
        data_elem = {k: 0 for k in src_data}
        # Результирующий словарь счетчиков
        data_counter = {
            self.SOCIAL_NETS: {
                v['id']: data_elem.copy()
                for v in self.social_nets_list
            },
            self.SOURCES: {
                v['id']: data_elem.copy()
                for v in self.sources_list
            },
        }
        data_counter.update({
            self.OTHERS: {
                self.other_social_net: data_counter[self.SOCIAL_NETS].pop(
                    self.other_social_net
                ),
            },
        })
        return data_counter

    def __generate_data(self, data_counter: dict, src_data: dict) -> dict:
        data = {}

        for key, _data_counter in data_counter.items():
            data[key] = []
            _src_data = self.__get_src_data(key)

            for _id, _data in _data_counter.items():
                data[key].append({
                    'id': _id,
                    'name': _src_data[_id],
                    self.STATUSES: [
                        {'id': k, 'name': src_data[k], self.COUNT: v}
                        for k, v in _data.items()
                    ],
                })

        return data

    ##################### Генераторы статистических данных #####################
    async def _generate_general_statistic(self) -> dict:
        # Результирующий список данных:
        data = []

        # Получаем список данных
        await self.get_obj()
        # Получаем список социальных сетей
        await self.get_social_nets()

        # Подготавливаем результирующий словарь счетчиков
        data_counter = self.__get_data_counter(ObjStatus.ALL)

        # Заполняем словарь счетчиков
        for o in self.obj_list:
            status_id = o[self.DB_STATUS]
            social_net_id = o[self.DB_SOCIAL_NET_ID]

            data_counter[self.SOCIAL_NETS][social_net_id][status_id] += 1

        data = self.__generate_data(data_counter, ObjStatus.ALL)
        return data
    """

    async def generate_statistic(self, statistic_type: str = None) -> dict:
        statistic = statistic_type or self.statistic_type

        if statistic == self.GENERAL:
            self._data = await self._generate_general_statistic()

        return self._data
