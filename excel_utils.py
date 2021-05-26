"""Утилиты для генерации excel-таблиц с помощью xlsxwriter."""
from openpyxl.styles import (
    Alignment as PyxlAlignment,
    Border as PyxlBorder,
    Font as PyxlFont,
    Protection as PyxlProtection,
    PatternFill as PyxlPattern,
    Side as PyxlSide,
    NamedStyle as PyxlStyle,
)
from openpyxl.utils.exceptions import ReadOnlyWorkbookException
from openpyxl.workbook import Workbook as PyxlWorkbook
from openpyxl.worksheet._write_only import WriteOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet as PyxlWorksheet


class Worksheet(PyxlWorksheet):
    def cell(self, row, column, value=None, style: PyxlStyle = None):
        """
        Returns a cell object based on the given coordinates.

        Usage: cell(row=15, column=1, value=5)

        Calling `cell` creates cells in memory when they
        are first accessed.

        :param row: row index of the cell (e.g. 4)
        :type row: int

        :param column: column index of the cell (e.g. 3)
        :type column: int

        :param value: value of the cell (e.g. 5)
        :type value: numeric or time or string or bool or none

        :rtype: openpyxl.cell.cell.Cell
        """

        if row < 1 or column < 1:
            raise ValueError("Row or column values must be at least 1")

        cell = self._get_cell(row, column)
        if value is not None:
            cell.value = value

        if style is not None:
            cell.style = style

        return cell


class Workbook(PyxlWorkbook):
    def create_sheet(self, title=None, index=None):
        """Create a worksheet (at an optional index).

        :param title: optional title of the sheet
        :type title: str
        :param index: optional position at which the sheet will be inserted
        :type index: int

        """
        if self.read_only:
            raise ReadOnlyWorkbookException(
                'Cannot create new sheet in a read-only workbook'
            )

        if self.write_only:
            new_ws = WriteOnlyWorksheet(parent=self, title=title)
        else:
            new_ws = Worksheet(parent=self, title=title)

        self._add_sheet(sheet=new_ws, index=index)
        return new_ws


class Style:
    """Класс стилей excel-таблицы.

    Args:
        style_name (str):                       Имя создаваемого объекта Стиль.
        style (optional, PyxlStyle):            Стиль, подлежащий изменению.
        alignment (optional, PyxlAlignment):    Выравнивание,
                                                подлежащee изменению.
        borders (optional, PyxlBorder):         Стиль границ рамки,
                                                подлежащий изменению.
        font (optional, PyxlFont):              Шрифт, подлежащий изменению.
        pattern (optional, PyxlPattern):        Шаблон заливки,
                                                подлежащий изменению.
        number_format (optional, str):          Формат чисел в ячейке.

    """
    class Alignment:
        # Horizontal
        GENERAL             = 'general'
        LEFT                = 'left'
        CENTER              = 'center'
        RIGHT               = 'right'
        FILL                = 'fill'
        JUSTIFY             = 'justify'
        CENTER_CONTINUOUS   = 'centerContinuous'
        DISTRIBUTED         = 'distributed'
        # Vertical
        TOP                 = 'top'
        VCENTER             = 'center'
        BOTTOM              = 'bottom'
        VJUSTIFY            = 'justify'
        VDISTRIBUTED        = 'distributed'

    class Border:
        NONE                = 'None'
        DASHDOT             = 'dashDot'
        DASHDOTDOT          = 'dashDotDot'
        DASHED              = 'dashed'
        DOTTED              = 'dotted'
        DOUBLE              = 'double'
        HAIR                = 'hair'
        MEDIUM              = 'medium'
        MEDIUMDASHDOT       = 'mediumDashDot'
        MEDIUMDASHDOTDOT    = 'mediumDashDotDot'
        MEDIUMDASHED        = 'mediumDashed'
        SLANTDASHDOT        = 'slantDashDot'
        THICK               = 'thick'
        THIN                = 'thin'

    class Font:
        ARIAL               = 'Arial'
        CALIBRI             = 'Calibri'
        TIMES_NEW_ROMAN     = 'Times New Roman'

    DEFAULT_FONT_SIZE       = 14

    class Format:
        GENERAL                     = 'General'
        TEXT                        = '@'
        NUMBER                      = '0'
        NUMBER_00                   = '0.00'
        NUMBER_COMMA_SEPARATED      = '#,##0'
        NUMBER_COMMA_SEPARATED_00   = '#,##0.00'
        PERCENTAGE                  = '0%'
        PERCENTAGE_00               = '0.00%'
        DATE_YYYYMMDD2              = 'yyyy-mm-dd'
        DATE_YYMMDD                 = 'yy-mm-dd'
        DATE_DDMMYY                 = 'dd/mm/yy'
        DATE_DMYSLASH               = 'd/m/y'
        DATE_DMYMINUS               = 'd-m-y'
        DATE_DMMINUS                = 'd-m'
        DATE_MYMINUS                = 'm-y'
        DATE_MMDDYY                 = 'mm-dd-yy'
        DATE_DATETIME               = 'yyyy-mm-dd h:mm:ss'
        DATE_HM_                    = 'h:mm AM/PM'
        DATE_HMS_                   = 'h:mm:ss AM/PM'
        DATE_HM                     = 'h:mm'
        DATE_MS                     = 'mm:ss'
        DATE_HMS                    = 'h:mm:ss'
        DATE_TIMEDELTA              = '[hh]:mm:ss'
        CURRENCY_USD_00             = '"$"#,##0.00_-'
        CURRENCY_USD                = '$#,##0_-'
        CURRENCY_EUR_00             = '[$EUR ]#,##0.00_-'
        CURRENCY_EUR                = '[$EUR ]#,##0_-'

    class Pattern:
        NONE                        = 'none'
        SOLID                       = 'solid'
        PATTERN_DARKDOWN            = 'darkDown'
        PATTERN_DARKGRAY            = 'darkGray'
        PATTERN_DARKGRID            = 'darkGrid'
        PATTERN_DARKHORIZONTAL      = 'darkHorizontal'
        PATTERN_DARKTRELLIS         = 'darkTrellis'
        PATTERN_DARKUP              = 'darkUp'
        PATTERN_DARKVERTICAL        = 'darkVertical'
        PATTERN_GRAY0625            = 'gray0625'
        PATTERN_GRAY125             = 'gray125'
        PATTERN_LIGHTDOWN           = 'lightDown'
        PATTERN_LIGHTGRAY           = 'lightGray'
        PATTERN_LIGHTGRID           = 'lightGrid'
        PATTERN_LIGHTHORIZONTAL     = 'lightHorizontal'
        PATTERN_LIGHTTRELLIS        = 'lightTrellis'
        PATTERN_LIGHTUP             = 'lightUp'
        PATTERN_LIGHTVERTICAL       = 'lightVertical'
        PATTERN_MEDIUMGRAY          = 'mediumGray'

    class Color:
        NONE_BLACK                  = 'FF000000'
        NONE_WHITE                  = 'FFFFFFFF'
        MAROON   	                = '00800000'
        DARK_RED	                = '008B0000'
        BROWN	                    = '00A52A2A'
        FIREBRICK	                = '00B22222'
        CRIMSON	                    = '00DC143C'
        RED	                        = '00FF0000'
        TOMATO                      = '00FF6347'
        CORAL	                    = '00FF7F50'
        INDIAN_RED	                = '00CD5C5C'
        LIGHT_CORAL	                = '00F08080'
        DARK_SALMON	                = '00E9967A'
        SALMON	                    = '00FA8072'
        LIGHT_SALMON	            = '00FFA07A'
        ORANGE_RED	                = '00FF4500'
        DARK_ORANGE	                = '00FF8C00'
        ORANGE      	            = '00FFA500'
        GOLD	                    = '00FFD700'
        DARK_GOLDEN_ROD	            = '00B8860B'
        GOLDEN_ROD	                = '00DAA520'
        PALE_GOLDEN_ROD	            = '00EEE8AA'
        DARK_KHAKI	                = '00BDB76B'
        KHAKI	                    = '00F0E68C'
        OLIVE	                    = '00808000'
        YELLOW	                    = '00FFFF00'
        YELLOW_GREEN	            = '009ACD32'
        DARK_OLIVE_GREEN            = '00556B2F'
        OLIVE_DRAB	                = '006B8E23'
        LAWN_GREEN	                = '007CFC00'
        CHART_REUSE	                = '007FFF00'
        GREEN_YELLOW	            = '00ADFF2F'
        DARK_GREEN	                = '00006400'
        GREEN	                    = '00008000'
        FOREST_GREEN	            = '00228B22'
        LIME	                    = '0000FF00'
        LIME_GREEN	                = '0032CD32'
        LIGHT_GREEN	                = '0090EE90'
        PALE_GREEN	                = '0098FB98'
        DARK_SEA_GREEN	            = '008FBC8F'
        MEDIUM_SPRING_GREEN	        = '0000FA9A'
        SPRING_GREEN	            = '0000FF7F'
        SEA_GREEN	                = '002E8B57'
        MEDIUM_AQUA_MARINE	        = '0066CDAA'
        MEDIUM_SEA_GREEN	        = '003CB371'
        LIGHT_SEA_GREEN	            = '0020B2AA'
        DARK_SLATE_GRAY	            = '002F4F4F'
        TEAL	                    = '00008080'
        DARK_CYAN	                = '00008B8B'
        AQUA	                    = '0000FFFF'
        CYAN	                    = '0000FFFF'
        LIGHT_CYAN	                = '00E0FFFF'
        DARK_TURQUOISE	            = '0000CED1'
        TURQUOISE	                = '0040E0D0'
        PALE_TURQUOISE	            = '00AFEEEE'
        AQUA_MARINE	                = '007FFFD4'
        POWDER_BLUE	                = '00B0E0E6'
        CADET_BLUE	                = '005F9EA0'
        STEEL_BLUE	                = '004682B4'
        CORN_FLOWER_BLUE	        = '006495ED'
        DEEP_SKY_BLUE	            = '0000BFFF'
        DODGER_BLUE	                = '001E90FF'
        LIGHT_BLUE	                = '00ADD8E6'
        SKY_BLUE	                = '0087CEEB'
        LIGHT_SKY_BLUE	            = '0087CEFA'
        MIDNIGHT_BLUE	            = '00191970'
        NAVY	                    = '00000080'
        DARK_BLUE	                = '0000008B'
        MEDIUM_BLUE	                = '000000CD'
        BLUE	                    = '000000FF'
        ROYAL_BLUE	                = '004169E1'
        BLUE_VIOLET	                = '008A2BE2'
        INDIGO	                    = '004B0082'
        DARK_SLATE_BLUE	            = '00483D8B'
        SLATE_BLUE	                = '006A5ACD'
        MEDIUM_SLATE_BLUE	        = '007B68EE'
        MEDIUM_PURPLE	            = '009370DB'
        DARK_MAGENTA	            = '008B008B'
        DARK_VIOLET	                = '009400D3'
        DARK_ORCHID	                = '009932CC'
        MEDIUM_ORCHID	            = '00BA55D3'
        PURPLE	                    = '00800080'
        THISTLE	                    = '00D8BFD8'
        PLUM	                    = '00DDA0DD'
        VIOLET	                    = '00EE82EE'
        MAGENTA	                    = '00FF00FF'
        ORCHID	                    = '00DA70D6'
        MEDIUM_VIOLET_RED	        = '00C71585'
        PALE_VIOLET_RED	            = '00DB7093'
        DEEP_PINK	                = '00FF1493'
        HOT_PINK	                = '00FF69B4'
        LIGHT_PINK	                = '00FFB6C1'
        PINK	                    = '00FFC0CB'
        ANTIQUE_WHITE	            = '00FAEBD7'
        BEIGE	                    = '00F5F5DC'
        BISQUE	                    = '00FFE4C4'
        BLANCHED_ALMOND	            = '00FFEBCD'
        WHEAT	                    = '00F5DEB3'
        CORN_SILK	                = '00FFF8DC'
        LEMON_CHIFFON	            = '00FFFACD'
        LIGHT_GOLDEN_ROD_YELLOW     = '00FAFAD2'
        LIGHT_YELLOW	            = '00FFFFE0'
        SADDLE_BROWN	            = '008B4513'
        SIENNA	                    = '00A0522D'
        CHOCOLATE	                = '00D2691E'
        PERU	                    = '00CD853F'
        SANDY_BROWN	                = '00F4A460'
        BURLY_WOOD	                = '00DEB887'
        TAN	                        = '00D2B48C'
        ROSY_BROWN	                = '00BC8F8F'
        MOCCASIN	                = '00FFE4B5'
        NAVAJO_WHITE	            = '00FFDEAD'
        PEACH_PUFF	                = '00FFDAB9'
        MISTY_ROSE	                = '00FFE4E1'
        LAVENDER_BLUSH	            = '00FFF0F5'
        LINEN	                    = '00FAF0E6'
        OLD_LACE	                = '00FDF5E6'
        PAPAYA_WHIP	                = '00FFEFD5'
        SEA_SHELL	                = '00FFF5EE'
        MINT_CREAM	                = '00F5FFFA'
        SLATE_GRAY	                = '00708090'
        LIGHT_SLATE_GRAY	        = '00778899'
        LIGHT_STEEL_BLUE	        = '00B0C4DE'
        LAVENDER	                = '00E6E6FA'
        FLORAL_WHITE	            = '00FFFAF0'
        ALICE_BLUE	                = '00F0F8FF'
        GHOST_WHITE	                = '00F8F8FF'
        HONEYDEW	                = '00F0FFF0'
        IVORY	                    = '00FFFFF0'
        AZURE	                    = '00F0FFFF'
        SNOW	                    = '00FFFAFA'
        BLACK	                    = '00000000'
        DIM_GRAY    	            = '00696969'
        GRAY	                    = '00808080'
        DARK_GRAY   	            = '00A9A9A9'
        SILVER	                    = '00C0C0C0'
        LIGHT_GRAY  	            = '00D3D3D3'
        GAINSBORO   	            = '00DCDCDC'
        WHITE_SMOKE 	            = '00F5F5F5'
        WHITE                       = '00FFFFFF'

    # # Множитель размера шрифта по вертикали.
    FONT_HEIGHT_INC = 2
    # Множитель размера шрифта по горизонтали.
    FONT_WIDTH_INC = 1

    def __init__(
        self,
        style_name:     str,
        style:          PyxlStyle = None,
        font:           PyxlFont = None,
        alignment:      PyxlAlignment = None,
        border:         PyxlBorder = None,
        pattern:        PyxlPattern = None,
        number_format:  str = None,
        **kwargs
    ):
        self.style_name         = style_name
        self.protection         = PyxlProtection()
        # Если стиля не существует, создаём
        if style is None:
            self.alignment      = self.get_alignment(alignment, **kwargs)
            self.border         = self.get_borders(border, **kwargs)
            self.font           = self.get_font(font, **kwargs)
            self.pattern        = self.get_pattern(pattern, **kwargs)
            self.number_format  = number_format or self.Format.GENERAL
        # Иначе - создаем копию стиля с изменением нужных параметров
        else:
            self.alignment      = alignment or self.get_alignment(
                                                    style.alignment, **kwargs)
            self.border         = border  or self.get_borders(style.border,
                                                                    **kwargs)
            self.font           = font      or self.get_font(style.font,
                                                                    **kwargs)
            self.pattern        = pattern   or self.get_pattern(style.fill,
                                                                    **kwargs)
            self.number_format  = number_format or style.number_format

    def get_alignment(self, alignment: PyxlAlignment=None,
                            **kwargs) -> PyxlAlignment:
        """Установка выравнивания содержимого ячеек.

        Args:
            alignment (optional, PyxlAlignment):Выравнивание,
                                                подлежащee изменению.
            **kwargs:
                align_v (optional, int):        Выравнивание по вертикали
                                                (по умолчанию - VCENTER).
                align_h (optional, int):        Выравнивание по горизонтали
                                                (по умолчанию - CENTER).
                align_rotate (optional, bool):  Поворот текста на 180*
                                                (по умолчанию - False).
                align_wrap (optional, bool):    Перенос многострочного текста
                                                (по умолчанию - False).
                align_shrink (optional, bool):  Размер ячейки по содержимому
                                                (по умолчанию - False).

        Returns:
            PyxlAlignment:  Объект 'Выравнивание' библиотеки openpyxl.

        """
        new_alignment = PyxlAlignment()

        # Если объект не существует, создаём
        if alignment is None:
            new_alignment.vertical = kwargs.get('align_v',
                                                    self.Alignment.VCENTER)
            new_alignment.horizontal = kwargs.get('align_h',
                                                    self.Alignment.CENTER)
            new_alignment.text_rotation = kwargs.get('align_rotate', False)
            new_alignment.wrap_text = kwargs.get('align_wrap', False)
            new_alignment.shrink_to_fit = kwargs.get('align_shrink', False)
        # Иначе - создаем копию объекта с изменением нужных параметров
        else:
            new_alignment.vertical = kwargs.get('align_v',
                                                    alignment.vertical)
            new_alignment.horizontal = kwargs.get('align_h',
                                                    alignment.horizontal)
            new_alignment.text_rotation = kwargs.get('align_rotate',
                                                    alignment.text_rotation)
            new_alignment.wrap_text = kwargs.get('align_wrap',
                                                    alignment.wrap_text)
            new_alignment.shrink_to_fit = kwargs.get('align_shrink',
                                                    alignment.shrink_to_fit)

        return new_alignment

    def get_borders(self, border: PyxlBorder=None, **kwargs) -> PyxlBorder:
        """Установка стиля границ рамок ячеек.

        Args:
            borders (optional, PyxlBorder):     Стиль границ рамки,
                                                подлежащий изменению.
            **kwargs:
                border_b (optional, int):       Нижняя граница рамки
                                                (по умолчанию - без рамки).
                border_l (optional, int):       Левая граница рамки
                                                (по умолчанию - без рамки).
                border_r (optional, int):       Правая граница рамки
                                                (по умолчанию - без рамки).
                border_t (optional, int):       Верхняя граница рамки
                                                (по умолчанию - без рамки).
                border_color_b (optional, str): Цвет нижней границы рамки
                                                (по умолчанию - DARK_GRAY).
                border_color_l (optional, str): Цвет левой границы рамки
                                                (по умолчанию - DARK_GRAY).
                border_color_r (optional, str): Цвет правой границы рамки
                                                (по умолчанию - DARK_GRAY).
                border_color_t (optional, str): Цвет верхней границы рамки
                                                (по умолчанию - DARK_GRAY).

        Returns:
            PyxlBorder:     Объект 'Границы' библиотеки openpyxl.

        """
        new_border = PyxlBorder()

        # Если объект не существует, создаём
        if border is None:
            new_border.bottom = PyxlSide(
                kwargs.get('border_b', self.Border.NONE),
                kwargs.get('border_color_b', self.Color.BLACK)
            )
            new_border.left = PyxlSide(
                kwargs.get('border_l', self.Border.NONE),
                kwargs.get('border_color_l', self.Color.BLACK)
            )
            new_border.right = PyxlSide(
                kwargs.get('border_r', self.Border.NONE),
                kwargs.get('border_color_r', self.Color.BLACK)
            )
            new_border.top = PyxlSide(
                kwargs.get('border_t', self.Border.NONE),
                kwargs.get('border_color_t', self.Color.BLACK)
            )
        # Иначе - создаем копию объекта с изменением нужных параметров
        else:
            new_border.bottom = PyxlSide(
                kwargs.get('border_b', border.bottom.style),
                kwargs.get('border_color_b', border.bottom.color)
            )
            new_border.left = PyxlSide(
                kwargs.get('border_l', border.left.style),
                kwargs.get('border_color_l', border.left.color)
            )
            new_border.right = PyxlSide(
                kwargs.get('border_r', border.right.style),
                kwargs.get('border_color_r', border.right.color)
            )
            new_border.top = PyxlSide(
                kwargs.get('border_t', border.top.style),
                kwargs.get('border_color_t', border.top.color)
            )

        return new_border

    def get_font(self, font: PyxlFont=None, **kwargs) -> PyxlFont:
        """Установка шрифта.

        Args:
            font (optional, PyxlFont):      Шрифт, подлежащий изменению.
            **kwargs:
                font_name (optional, str):  Наименование шрифта
                                            (по умолчанию - TIMES_NEW_ROMAN).
                font_size (optional, int):  Размер шрифта
                                            (по умолчанию - DEFAULT_FONT_SIZE).
                font_color (optional, str): Цвет шрифта
                                            (по умолчанию - BLACK).
                font_bold (optional, bool): Жирный шрифт
                                            (по умолчанию - False = выключен).
                font_italic (optional,
                                bool):      Курсив
                                            (по умолчанию - False = выключен).

        Returns:
            PyxlFont:   Объект 'Шрифт' библиотеки openpyxl.

        """
        new_font = PyxlFont()

        # Если объект не существует, создаём
        if font is None:
            new_font.name = kwargs.get('font_name', self.Font.TIMES_NEW_ROMAN)
            new_font.size = kwargs.get('font_size', self.DEFAULT_FONT_SIZE)
            new_font.color = kwargs.get('font_color', self.Color.BLACK)
            new_font.bold = kwargs.get('font_bold', False)
            new_font.italic = kwargs.get('font_italic', False)
        # Иначе - создаем копию объекта с изменением нужных параметров
        else:
            new_font.name = kwargs.get('font_name', font.name)
            new_font.size = kwargs.get('font_size', font.size)
            new_font.color = kwargs.get('font_color', font.color)
            new_font.bold = kwargs.get('font_bold', font.bold)
            new_font.italic = kwargs.get('font_italic', font.italic)

        return new_font

    def get_pattern(self, pattern: PyxlPattern=None, **kwargs) -> PyxlPattern:
        """Установка шаблона заливки ячеек.

        Args:
            pattern (optional, PyxlPattern):    Шаблон заливки,
                                                подлежащий изменению.
            **kwargs:
                pattern_type (optional, int):   Выравнивание по горизонтали
                                                (по умолчанию - NO_PATTERN).
                pattern_fg_color (optional,
                                        str):   Стиль заливки (фронт)
                                                (по умолчанию - NONE_WHITE).
                pattern_bg_color (optional,
                                        str):   Стиль заливки (бэк)
                                                (по умолчанию - NONE_BLACK).

        Returns:
            PyxlPattern:  Объект 'Заливка' библиотеки openpyxl.

        """
        new_pattern = PyxlPattern()

        # Если объект не существует, создаём
        if pattern is None:
            new_pattern.fill_type = kwargs.get('pattern_type',
                                                        self.Pattern.NONE)
            new_pattern.start_color = kwargs.get('pattern_fg_color',
                                                        self.Color.NONE_WHITE)
            new_pattern.end_color = kwargs.get('pattern_bg_color',
                                                        self.Color.NONE_BLACK)
        # Иначе - создаем копию объекта с изменением нужных параметров
        else:
            new_pattern.fill_type = kwargs.get('pattern_type',
                                                        pattern.fill_type)
            new_pattern.start_color = kwargs.get('pattern_fg_color',
                                                        pattern.start_color)
            new_pattern.end_color = kwargs.get('pattern_bg_color',
                                                        pattern.end_color)

        return new_pattern

    def get_style(self) -> PyxlStyle:
        """Создает новый объект стиля ячейки с заданными в классе параметрами.

        Returns:
            PyxlStyle:  Объект 'Стиль' библиотеки openpyxl.
        """
        return PyxlStyle(
            name            = self.style_name,
            alignment       = self.alignment,
            border          = self.border,
            font            = self.font,
            fill            = self.pattern,
            number_format   = self.number_format,
            protection      = self.protection
        )
